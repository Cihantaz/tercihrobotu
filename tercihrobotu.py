import io
import json
import os
import sqlite3
import threading
import time
import uuid
import zlib
from datetime import datetime, timedelta, timezone
from functools import wraps
from pathlib import Path

from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

import pandas as pd
from flask import (
    Flask,
    abort,
    flash,
    g,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.security import check_password_hash

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "exam_secret_key")
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)

APP_ROOT = Path(app.root_path)
INSTANCE_DIR = APP_ROOT / "instance"
INSTANCE_DIR.mkdir(exist_ok=True)
DEFAULT_DATA_FILE = Path("data") / "23_24_isik.xlsx"
DATABASE_PATH = Path(os.environ.get("DATABASE_PATH", str(INSTANCE_DIR / "tercihrobotu.db")))
if not DATABASE_PATH.is_absolute():
    DATABASE_PATH = APP_ROOT / DATABASE_PATH
REPORT_RETENTION_DAYS = int(os.environ.get("REPORT_RETENTION_DAYS", "30"))
LOG_RETENTION_DAYS = int(os.environ.get("LOG_RETENTION_DAYS", "60"))
MAX_PARAMETER_COUNT = int(os.environ.get("MAX_PARAMETER_COUNT", "12"))
ADMIN_USERNAME = os.environ.get("ADMIN_USERNAME", "cihan.tazeoz@isikun.edu.tr")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "11235813")
ADMIN_PASSWORD_HASH = os.environ.get("ADMIN_PASSWORD_HASH", "")
SUPPORTED_LANGS = ("tr", "en")

TEXTS = {
    "tr": {
        "page_title": "Tercih Robotu",
        "brand_title": "Tercih Robotu",
        "login_title": "Tercih Robotu",
        "login_prompt": "Devam etmek için mail adresinizi ve telefon numaranızı giriniz.",
        "language_label": "Dil",
        "language_tr": "TR",
        "language_en": "EN",
        "email_label": "Mail adresinizi giriniz",
        "email_placeholder": "ornek@mail.com",
        "phone_label": "Telefon numaranızı giriniz",
        "phone_placeholder": "05xx xxx xx xx",
        "privacy_consent_text": "Tercih Robotu sistemimiz, adayların tercih süreçlerine rehberlik etmek amacıyla tasarlanmış olup tavsiye niteliği taşımaktadır. Sistem sonuçları, üniversiteye kesin yerleşme veya burs kazanma garantisi sunmamaktadır.",
        "privacy_consent_prefix": "KVKK kapsamında kişisel verilerimin işlenmesine ilişkin",
        "privacy_consent_link_text": "Aydınlatma Metni'ni",
        "privacy_consent_suffix": "okudum ve kabul ediyorum.",
        "info_consent_label": "Yukarıdaki bilgilendirme metnini okudum ve onaylıyorum.",
        "continue_button": "Devam Et",
        "home_button": "Ana Sayfa",
        "guide_button": "Kullanım Kılavuzu",
        "admin_button": "Admin",
        "form_intro": "Lütfen aşağıdaki bilgileri doldurunuz.",
        "student_name_label": "Öğrenci Adı Soyadı ve gerekirse bölüm",
        "student_name_placeholder": "Örnek: Ayşe Yılmaz, Psikoloji",
        "ranking_label": "Sıralama",
        "score_type_label": "Puan Türü",
        "limit_label": "Sınır",
        "scenario_add": "Senaryo Ekle",
        "scenario_title": "Senaryolar",
        "delete_button": "Sil",
        "analyze_button": "Analiz Et",
        "results_title": "Analiz Sonuçları",
        "download_button": "PDF Olarak İndir",
        "download_pdf_button": "PDF Olarak İndir",
        "no_results": "Bu senaryolar için sonuç bulunamadı.",
        "footer_text": "Işık Üniversitesi Öğrenci İşleri Daire Başkanlığı",
        "fill_required": "Sıralama, Puan Türü ve Sınır alanlarını doldurun.",
        "invalid_email": "Geçerli bir mail adresi girin.",
        "invalid_phone": "Telefon numarası girin.",
        "invalid_privacy_consent": "Aydınlatma Metni onay kutusunu işaretleyin.",
        "invalid_info_consent": "Bilgilendirme metni onay kutusunu işaretleyin.",
        "invalid_session": "Giriş bilgileri geçerli değil, tekrar girin.",
        "invalid_scenarios": "Senaryo listesi okunamadı.",
        "no_scenario": "En az bir geçerli senaryo ekleyin.",
        "too_many_scenarios": "En fazla {max_count} senaryo işlenir. İlk {max_count} senaryo kullanıldı.",
        "data_file_error": "Veri dosyası okunamadı: {error}",
        "no_result_flash": "Sonuç bulunamadı.",
        "table_program": "Bölüm Adı",
        "table_score_type": "Puan Türü",
        "table_scholarship": "Burs Oranı",
        "table_base_ranking": "Taban Sıralama",
        "table_base_score": "Taban Puan",
        "table_ceiling_score": "Tavan Puan",
        "table_fee": "Ücret",
        "table_language": "Dil",
        "table_quota": "Kontenjan",
        "table_status": "Etiket",
        "table_parameter": "Parametre",
        "status_eligible": "Uygun",
        "status_risky": "Riskli",
        "status_out": "Uygunsuz",
        "status_unknown": "Sıralama Verisi Yok",
        "sheet_name": "Sonuçlar",
        "excel_student": "Öğrenci",
        "excel_department": "Talep Edilen Bölüm",
        "excel_report": "Rapor No",
    },
    "en": {
        "page_title": "Preference Robot",
        "brand_title": "Tercih Robotu",
        "login_title": "Preference Robot",
        "login_prompt": "Enter your email address and phone number to continue.",
        "language_label": "Language",
        "language_tr": "TR",
        "language_en": "EN",
        "email_label": "Enter your email address",
        "email_placeholder": "example@mail.com",
        "phone_label": "Enter your phone number",
        "phone_placeholder": "+90 5xx xxx xx xx",
        "privacy_consent_text": "Our Preference Robot system is designed to guide candidates in their preference process and is provided for informational purposes only. System results do not guarantee university admission or scholarship awards.",
        "privacy_consent_prefix": "I have read and accept the",
        "privacy_consent_link_text": "Clarification Text",
        "privacy_consent_suffix": "regarding the processing of my personal data under KVKK.",
        "info_consent_label": "I have read and confirm the above notice.",
        "continue_button": "Continue",
        "home_button": "Home",
        "guide_button": "User Guide",
        "admin_button": "Admin",
        "form_intro": "Please fill in the information below.",
        "student_name_label": "Student full name and department if needed",
        "student_name_placeholder": "Example: Ayse Yilmaz, Psychology",
        "ranking_label": "Ranking",
        "score_type_label": "Score Type",
        "limit_label": "Limit",
        "scenario_add": "Add Scenario",
        "scenario_title": "Scenarios",
        "delete_button": "Delete",
        "analyze_button": "Analyze",
        "results_title": "Analysis Results",
        "download_button": "Download as PDF",
        "download_pdf_button": "Download as PDF",
        "no_results": "No results were found for these scenarios.",
        "footer_text": "Isik University Registrar's Office",
        "fill_required": "Fill in the Ranking, Score Type and Limit fields.",
        "invalid_email": "Enter a valid email address.",
        "invalid_phone": "Enter a phone number.",
        "invalid_privacy_consent": "Select the clarification text consent checkbox.",
        "invalid_info_consent": "Please confirm the information notice checkbox.",
        "invalid_session": "Login information is invalid. Please start again.",
        "invalid_scenarios": "Scenario list could not be read.",
        "no_scenario": "Add at least one valid scenario.",
        "too_many_scenarios": "A maximum of {max_count} scenarios can be processed. Only the first {max_count} scenarios were used.",
        "data_file_error": "The data file could not be read: {error}",
        "no_result_flash": "No results were found.",
        "table_program": "Program Name",
        "table_score_type": "Score Type",
        "table_scholarship": "Scholarship",
        "table_base_ranking": "Base Ranking",
        "table_base_score": "Base Score",
        "table_ceiling_score": "Top Score",
        "table_fee": "Tuition",
        "table_language": "Language",
        "table_quota": "Quota",
        "table_status": "Status",
        "table_parameter": "Parameter",
        "status_eligible": "Eligible",
        "status_risky": "Stretch",
        "status_out": "Not Listed",
        "status_unknown": "Ranking Data Missing",
        "sheet_name": "Results",
        "excel_student": "Student",
        "excel_department": "Requested Department",
        "excel_report": "Report ID",
    },
}

TABLE_HEADER_KEYS = [
    ("bolum_adi", "table_program"),
    ("puan_turu", "table_score_type"),
    ("burs_orani", "table_scholarship"),
    ("taban_siralama", "table_base_ranking"),
    ("taban_puan", "table_base_score"),
    ("tavan_puan", "table_ceiling_score"),
    ("ucret", "table_fee"),
    ("dil", "table_language"),
    ("kontenjan", "table_quota"),
    ("etiket", "table_status"),
]

BURSLULUK_KELIMELERI = [
    "Burslu",
    "\u00dccretli",
    "%50 \u0130ndirimli",
    "%25 \u0130ndirimli",
    "%75 \u0130ndirimli",
    "%100 Burslu",
]

EN_BURS_MAP = {
    "Burslu": "Scholarship",
    "Ücretli": "Full Tuition",
    "%50 İndirimli": "50% Discount",
    "%25 İndirimli": "25% Discount",
    "%75 İndirimli": "75% Discount",
    "%100 Burslu": "100% Scholarship",
}

_dataset_cache = {"key": None, "data": None, "path": None, "loaded_at": None}
_dataset_lock = threading.Lock()
_cleanup_lock = threading.Lock()
_last_cleanup_ts = 0.0


def utcnow_iso():
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def normalize_lang(value):
    if value in SUPPORTED_LANGS:
        return value
    return "tr"


def get_texts(lang):
    return TEXTS[normalize_lang(lang)]


def get_table_headers(lang):
    texts = get_texts(lang)
    return [(key, texts[label_key]) for key, label_key in TABLE_HEADER_KEYS]


def find_system_font():
    candidates = [
        "DejaVuSans.ttf",
        "LiberationSans-Regular.ttf",
        "Arial.ttf",
        "FreeSans.ttf",
    ]
    search_paths = [
        Path("/usr/share/fonts/truetype/dejavu"),
        Path("/usr/share/fonts/truetype/liberation"),
        Path("/usr/share/fonts/truetype/freefont"),
        Path("/usr/share/fonts/truetype/msttcorefonts"),
        Path("/Library/Fonts"),
        Path("C:/Windows/Fonts"),
    ]
    for directory in search_paths:
        for candidate in candidates:
            path = directory / candidate
            if path.exists():
                return path
    return None


def get_pdf_font_name():
    font_name = "Helvetica"
    if font_name in pdfmetrics.getRegisteredFontNames():
        return font_name
    system_font = find_system_font()
    if system_font is not None:
        try:
            pdfmetrics.registerFont(TTFont("CustomFont", str(system_font)))
            return "CustomFont"
        except Exception:
            pass
    return font_name


def localize_status(status_key, lang):
    texts = get_texts(lang)
    key_map = {
        "eligible": "status_eligible",
        "risky": "status_risky",
        "out": "status_out",
        "unknown": "status_unknown",
    }
    return texts[key_map.get(status_key, "status_unknown")]


def translate_burs_orani(value, lang):
    if normalize_lang(lang) != "en":
        return value
    return EN_BURS_MAP.get(value, value)


def normalize_phone(value):
    return str(value or "").strip()


def is_valid_email(value):
    return bool(value and "@" in value)


def clean_filename(value):
    translation_table = str.maketrans(
        {
            "\u00e7": "c",
            "\u00c7": "C",
            "\u011f": "g",
            "\u011e": "G",
            "\u0131": "i",
            "\u0130": "I",
            "\u00f6": "o",
            "\u00d6": "O",
            "\u015f": "s",
            "\u015e": "S",
            "\u00fc": "u",
            "\u00dc": "U",
        }
    )
    sanitized = (value or "").translate(translation_table)
    safe_chars = []
    for char in sanitized:
        if char.isalnum() or char in {"-", "_"}:
            safe_chars.append(char)
        elif char in {" ", "."}:
            safe_chars.append("_")
    return "".join(safe_chars).strip("_") or "tercihrobotu_raporu"


def get_client_ip():
    forwarded_for = request.headers.get("X-Forwarded-For", "")
    if forwarded_for:
        return forwarded_for.split(",")[0].strip()
    return request.remote_addr or "unknown"


def get_user_agent():
    return request.headers.get("User-Agent", "unknown")[:500]


def get_db_connection():
    DATABASE_PATH.parent.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(str(DATABASE_PATH))
    connection.row_factory = sqlite3.Row
    try:
        connection.execute("PRAGMA journal_mode=WAL")
    except sqlite3.OperationalError:
        pass
    connection.execute("PRAGMA synchronous=NORMAL")
    connection.execute("PRAGMA foreign_keys=ON")
    return connection


def init_db():
    with get_db_connection() as connection:
        connection.executescript(
            """
            CREATE TABLE IF NOT EXISTS app_settings (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS analysis_runs (
                id TEXT PRIMARY KEY,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                student_email TEXT NOT NULL DEFAULT '',
                student_phone TEXT NOT NULL DEFAULT '',
                language TEXT NOT NULL DEFAULT 'tr',
                student_input TEXT NOT NULL DEFAULT '',
                student_name TEXT NOT NULL DEFAULT '',
                requested_department TEXT NOT NULL DEFAULT '',
                ranking_summary TEXT NOT NULL DEFAULT '',
                score_types_summary TEXT NOT NULL DEFAULT '',
                params_json TEXT NOT NULL,
                result_blob BLOB,
                result_count INTEGER NOT NULL DEFAULT 0,
                source_file TEXT NOT NULL,
                duration_ms INTEGER NOT NULL DEFAULT 0,
                status TEXT NOT NULL,
                error_message TEXT,
                client_ip TEXT,
                user_agent TEXT,
                download_count INTEGER NOT NULL DEFAULT 0
            );

            CREATE INDEX IF NOT EXISTS idx_analysis_created_at
            ON analysis_runs(created_at DESC);

            CREATE INDEX IF NOT EXISTS idx_analysis_status
            ON analysis_runs(status);

            CREATE TABLE IF NOT EXISTS download_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                analysis_id TEXT NOT NULL,
                created_at TEXT NOT NULL,
                filename TEXT NOT NULL,
                row_count INTEGER NOT NULL DEFAULT 0,
                client_ip TEXT,
                user_agent TEXT
            );

            CREATE INDEX IF NOT EXISTS idx_download_created_at
            ON download_events(created_at DESC);

            CREATE TABLE IF NOT EXISTS app_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                created_at TEXT NOT NULL,
                level TEXT NOT NULL,
                event_type TEXT NOT NULL,
                message TEXT NOT NULL,
                context_json TEXT
            );

            CREATE INDEX IF NOT EXISTS idx_logs_created_at
            ON app_logs(created_at DESC);

            CREATE TABLE IF NOT EXISTS student_profiles (
                student_email TEXT PRIMARY KEY,
                student_phone TEXT NOT NULL DEFAULT '',
                preferred_language TEXT NOT NULL DEFAULT 'tr',
                last_student_input TEXT NOT NULL DEFAULT '',
                last_student_name TEXT NOT NULL DEFAULT '',
                last_ranking_summary TEXT NOT NULL DEFAULT '',
                last_score_types_summary TEXT NOT NULL DEFAULT '',
                login_count INTEGER NOT NULL DEFAULT 0,
                analysis_count INTEGER NOT NULL DEFAULT 0,
                download_count INTEGER NOT NULL DEFAULT 0,
                guide_count INTEGER NOT NULL DEFAULT 0,
                event_count INTEGER NOT NULL DEFAULT 0,
                first_seen_at TEXT NOT NULL,
                last_seen_at TEXT NOT NULL,
                last_event_type TEXT NOT NULL DEFAULT '',
                last_analysis_id TEXT NOT NULL DEFAULT ''
            );

            CREATE TABLE IF NOT EXISTS student_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                created_at TEXT NOT NULL,
                student_email TEXT NOT NULL,
                student_phone TEXT NOT NULL DEFAULT '',
                language TEXT NOT NULL DEFAULT 'tr',
                event_type TEXT NOT NULL,
                student_input TEXT NOT NULL DEFAULT '',
                student_name TEXT NOT NULL DEFAULT '',
                ranking_summary TEXT NOT NULL DEFAULT '',
                score_types_summary TEXT NOT NULL DEFAULT '',
                analysis_id TEXT,
                status TEXT NOT NULL DEFAULT '',
                details_json TEXT
            );

            CREATE INDEX IF NOT EXISTS idx_student_events_created_at
            ON student_events(created_at DESC);

            CREATE INDEX IF NOT EXISTS idx_student_events_email
            ON student_events(student_email, created_at DESC);
            """
        )
        ensure_analysis_run_columns(connection)
        ensure_student_profile_columns(connection)
        ensure_student_event_columns(connection)


def ensure_analysis_run_columns(connection):
    columns = {
        row["name"]
        for row in connection.execute("PRAGMA table_info(analysis_runs)").fetchall()
    }

    if "student_email" not in columns:
        connection.execute(
            "ALTER TABLE analysis_runs ADD COLUMN student_email TEXT NOT NULL DEFAULT ''"
        )
    if "ranking_summary" not in columns:
        connection.execute(
            "ALTER TABLE analysis_runs ADD COLUMN ranking_summary TEXT NOT NULL DEFAULT ''"
        )
    if "student_phone" not in columns:
        connection.execute(
            "ALTER TABLE analysis_runs ADD COLUMN student_phone TEXT NOT NULL DEFAULT ''"
        )
    if "language" not in columns:
        connection.execute(
            "ALTER TABLE analysis_runs ADD COLUMN language TEXT NOT NULL DEFAULT 'tr'"
        )
    if "score_types_summary" not in columns:
        connection.execute(
            "ALTER TABLE analysis_runs ADD COLUMN score_types_summary TEXT NOT NULL DEFAULT ''"
        )


def ensure_student_profile_columns(connection):
    columns = {
        row["name"]
        for row in connection.execute("PRAGMA table_info(student_profiles)").fetchall()
    }
    required_columns = {
        "student_phone": "TEXT NOT NULL DEFAULT ''",
        "preferred_language": "TEXT NOT NULL DEFAULT 'tr'",
        "last_student_input": "TEXT NOT NULL DEFAULT ''",
        "last_student_name": "TEXT NOT NULL DEFAULT ''",
        "last_ranking_summary": "TEXT NOT NULL DEFAULT ''",
        "last_score_types_summary": "TEXT NOT NULL DEFAULT ''",
        "login_count": "INTEGER NOT NULL DEFAULT 0",
        "analysis_count": "INTEGER NOT NULL DEFAULT 0",
        "download_count": "INTEGER NOT NULL DEFAULT 0",
        "guide_count": "INTEGER NOT NULL DEFAULT 0",
        "event_count": "INTEGER NOT NULL DEFAULT 0",
        "first_seen_at": "TEXT NOT NULL DEFAULT ''",
        "last_seen_at": "TEXT NOT NULL DEFAULT ''",
        "last_event_type": "TEXT NOT NULL DEFAULT ''",
        "last_analysis_id": "TEXT NOT NULL DEFAULT ''",
    }
    for column_name, definition in required_columns.items():
        if column_name not in columns:
            connection.execute(
                "ALTER TABLE student_profiles ADD COLUMN {} {}".format(column_name, definition)
            )


def ensure_student_event_columns(connection):
    columns = {
        row["name"]
        for row in connection.execute("PRAGMA table_info(student_events)").fetchall()
    }
    required_columns = {
        "student_phone": "TEXT NOT NULL DEFAULT ''",
        "language": "TEXT NOT NULL DEFAULT 'tr'",
        "event_type": "TEXT NOT NULL DEFAULT ''",
        "student_input": "TEXT NOT NULL DEFAULT ''",
        "student_name": "TEXT NOT NULL DEFAULT ''",
        "ranking_summary": "TEXT NOT NULL DEFAULT ''",
        "score_types_summary": "TEXT NOT NULL DEFAULT ''",
        "analysis_id": "TEXT",
        "status": "TEXT NOT NULL DEFAULT ''",
        "details_json": "TEXT",
    }
    for column_name, definition in required_columns.items():
        if column_name not in columns:
            connection.execute(
                "ALTER TABLE student_events ADD COLUMN {} {}".format(column_name, definition)
            )


def maybe_cleanup(force=False):
    global _last_cleanup_ts

    now_ts = time.time()
    if not force and now_ts - _last_cleanup_ts < 3600:
        return

    with _cleanup_lock:
        if not force and now_ts - _last_cleanup_ts < 3600:
            return

        report_cutoff = (
            datetime.now(timezone.utc) - timedelta(days=REPORT_RETENTION_DAYS)
        ).replace(microsecond=0).isoformat().replace("+00:00", "Z")
        log_cutoff = (
            datetime.now(timezone.utc) - timedelta(days=LOG_RETENTION_DAYS)
        ).replace(microsecond=0).isoformat().replace("+00:00", "Z")

        with get_db_connection() as connection:
            connection.execute("DELETE FROM download_events WHERE created_at < ?", (report_cutoff,))
            connection.execute("DELETE FROM analysis_runs WHERE created_at < ?", (report_cutoff,))
            connection.execute("DELETE FROM app_logs WHERE created_at < ?", (log_cutoff,))
            connection.execute("DELETE FROM student_events WHERE created_at < ?", (log_cutoff,))

        _last_cleanup_ts = now_ts


def get_setting(key, default=None):
    with get_db_connection() as connection:
        row = connection.execute("SELECT value FROM app_settings WHERE key = ?", (key,)).fetchone()
    return row["value"] if row else default


def set_setting(key, value):
    timestamp = utcnow_iso()
    with get_db_connection() as connection:
        connection.execute(
            """
            INSERT INTO app_settings(key, value, updated_at)
            VALUES (?, ?, ?)
            ON CONFLICT(key) DO UPDATE SET
                value = excluded.value,
                updated_at = excluded.updated_at
            """,
            (key, value, timestamp),
        )


def log_event(level, event_type, message, context=None):
    payload = json.dumps(context or {}, ensure_ascii=False)
    timestamp = utcnow_iso()
    with get_db_connection() as connection:
        connection.execute(
            "INSERT INTO app_logs(created_at, level, event_type, message, context_json) VALUES (?, ?, ?, ?, ?)",
            (timestamp, level.upper(), event_type, message, payload),
        )

    app.logger.info("%s | %s | %s", level.upper(), event_type, message)


def record_student_event(
    student_email,
    event_type,
    *,
    student_phone="",
    language="tr",
    student_input="",
    student_name="",
    ranking_summary="",
    score_types_summary="",
    analysis_id="",
    status="success",
    details=None,
):
    student_email = str(student_email or "").strip().lower()
    if not is_valid_email(student_email):
        return

    timestamp = utcnow_iso()
    normalized_lang = normalize_lang(language)
    payload = json.dumps(details or {}, ensure_ascii=False)
    counters = {
        "login_count": 1 if event_type == "login" else 0,
        "analysis_count": 1 if event_type == "analysis" else 0,
        "download_count": 1 if event_type == "download" else 0,
        "guide_count": 1 if event_type == "guide_open" else 0,
        "event_count": 1,
    }

    with get_db_connection() as connection:
        connection.execute(
            """
            INSERT INTO student_events(
                created_at, student_email, student_phone, language, event_type,
                student_input, student_name, ranking_summary, score_types_summary,
                analysis_id, status, details_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                timestamp,
                student_email,
                normalize_phone(student_phone),
                normalized_lang,
                event_type,
                student_input,
                student_name,
                ranking_summary,
                score_types_summary,
                analysis_id or None,
                status,
                payload,
            ),
        )
        connection.execute(
            """
            INSERT INTO student_profiles(
                student_email, student_phone, preferred_language, last_student_input,
                last_student_name, last_ranking_summary, last_score_types_summary,
                login_count, analysis_count, download_count, guide_count, event_count,
                first_seen_at, last_seen_at, last_event_type, last_analysis_id
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(student_email) DO UPDATE SET
                student_phone = CASE
                    WHEN excluded.student_phone <> '' THEN excluded.student_phone
                    ELSE student_profiles.student_phone
                END,
                preferred_language = excluded.preferred_language,
                last_student_input = CASE
                    WHEN excluded.last_student_input <> '' THEN excluded.last_student_input
                    ELSE student_profiles.last_student_input
                END,
                last_student_name = CASE
                    WHEN excluded.last_student_name <> '' THEN excluded.last_student_name
                    ELSE student_profiles.last_student_name
                END,
                last_ranking_summary = CASE
                    WHEN excluded.last_ranking_summary <> '' THEN excluded.last_ranking_summary
                    ELSE student_profiles.last_ranking_summary
                END,
                last_score_types_summary = CASE
                    WHEN excluded.last_score_types_summary <> '' THEN excluded.last_score_types_summary
                    ELSE student_profiles.last_score_types_summary
                END,
                login_count = student_profiles.login_count + excluded.login_count,
                analysis_count = student_profiles.analysis_count + excluded.analysis_count,
                download_count = student_profiles.download_count + excluded.download_count,
                guide_count = student_profiles.guide_count + excluded.guide_count,
                event_count = student_profiles.event_count + excluded.event_count,
                last_seen_at = excluded.last_seen_at,
                last_event_type = excluded.last_event_type,
                last_analysis_id = CASE
                    WHEN excluded.last_analysis_id <> '' THEN excluded.last_analysis_id
                    ELSE student_profiles.last_analysis_id
                END
            """,
            (
                student_email,
                normalize_phone(student_phone),
                normalized_lang,
                student_input,
                student_name,
                ranking_summary,
                score_types_summary,
                counters["login_count"],
                counters["analysis_count"],
                counters["download_count"],
                counters["guide_count"],
                counters["event_count"],
                timestamp,
                timestamp,
                event_type,
                analysis_id,
            ),
        )


def admin_credentials_configured():
    return bool(ADMIN_USERNAME and (ADMIN_PASSWORD or ADMIN_PASSWORD_HASH))


def is_admin_authenticated():
    return session.get("is_admin") is True


def verify_admin_credentials(username, password):
    if not admin_credentials_configured():
        return False
    if username != ADMIN_USERNAME:
        return False
    if ADMIN_PASSWORD_HASH:
        return check_password_hash(ADMIN_PASSWORD_HASH, password)
    return password == ADMIN_PASSWORD


def admin_required(view_func):
    @wraps(view_func)
    def wrapped_view(*args, **kwargs):
        if not is_admin_authenticated():
            return redirect(url_for("admin_login", next=request.path))
        return view_func(*args, **kwargs)

    return wrapped_view


def resolve_repo_path(relative_path):
    candidate = (APP_ROOT / relative_path).resolve()
    candidate.relative_to(APP_ROOT.resolve())
    return candidate


def list_available_data_files():
    files = []
    data_dir = APP_ROOT / "data"
    if data_dir.exists():
        for path in sorted(data_dir.glob("*.xlsx")):
            files.append(path.relative_to(APP_ROOT).as_posix())
    for path in sorted(APP_ROOT.glob("*.xlsx")):
        files.append(path.relative_to(APP_ROOT).as_posix())
    unique_files = []
    seen = set()
    for item in files:
        if item.startswith("~$"):
            continue
        if item not in seen:
            seen.add(item)
            unique_files.append(item)
    return unique_files


def get_active_data_file_setting():
    configured = get_setting("active_data_file")
    if configured:
        return configured

    env_path = os.environ.get("DATA_FILE_PATH", "")
    if env_path:
        env_candidate = Path(env_path)
        if env_candidate.is_absolute():
            return str(env_candidate)
        return env_candidate.as_posix()

    return DEFAULT_DATA_FILE.as_posix()


def resolve_active_data_file():
    configured = get_active_data_file_setting()
    candidate = Path(configured)

    if candidate.is_absolute():
        absolute_path = candidate
        display_path = str(candidate)
    else:
        absolute_path = resolve_repo_path(candidate.as_posix())
        display_path = candidate.as_posix()

    if absolute_path.exists():
        return absolute_path, display_path

    available = list_available_data_files()
    if available:
        fallback = available[0]
        fallback_path = resolve_repo_path(fallback)
        if configured != fallback:
            set_setting("active_data_file", fallback)
        return fallback_path, fallback

    raise FileNotFoundError("Veri dosyasi bulunamadi.")


def clear_dataset_cache():
    with _dataset_lock:
        _dataset_cache["key"] = None
        _dataset_cache["data"] = None
        _dataset_cache["path"] = None
        _dataset_cache["loaded_at"] = None


def temizle_sayi(value):
    if value is None:
        return 0
    text_value = str(value).strip()
    if not text_value or text_value == "-":
        return 0
    text_value = text_value.replace(".", "").replace(",", "")
    digits = "".join(char for char in text_value if char.isdigit())
    if not digits:
        return 0
    try:
        return int(digits)
    except ValueError:
        return 0


def temizle_sayi_opsiyonel(value):
    if value is None:
        return None
    text_value = str(value).strip()
    if not text_value or text_value == "-":
        return None
    cleaned = temizle_sayi(text_value)
    return cleaned or None


def format_ucret(value):
    text_value = str(value).strip()
    if not text_value or text_value.lower() == "nan":
        return ""
    normalized = text_value.lower().replace("ç", "c").replace("ş", "s").replace("ı", "i").replace("ğ", "g").replace("ö", "o").replace("ü", "u")
    if "ucretsiz" in normalized or "free" in normalized:
        return "Ücretsiz" if "ucretsiz" in normalized else "Free"
    cleaned = text_value.replace(".", "").replace(",", "").replace("\u20ba", "").replace("TL", "").strip()
    if cleaned.isdigit():
        if int(cleaned) == 0:
            return "Ücretsiz"
        return "{:,.0f} TL".format(float(cleaned)).replace(",", ".")
    return text_value


def turkish_lower(value):
    if value is None:
        return ""
    return str(value).replace("İ", "i").replace("I", "ı").lower()


def infer_burs_orani(raw_burs, program_adi):
    burs = (raw_burs or "").strip()
    if burs:
        return burs
    lowered_program = turkish_lower(program_adi)
    for keyword in BURSLULUK_KELIMELERI:
        if turkish_lower(keyword) in lowered_program:
            return keyword
    return ""


def detect_language(program_adi, raw_dil=None):
    if raw_dil is not None:
        raw_value = str(raw_dil).strip()
        if raw_value:
            lowered = turkish_lower(raw_value)
            upper_value = raw_value.upper()
            if lowered in ("ingilizce", "en", "english", "ing") or upper_value in ("ING", "INGILIZCE", "EN", "ENGLISH"):
                return "EN"
            return "TR"
    lowered_program = turkish_lower(program_adi)
    return "EN" if "ingilizce" in lowered_program else "TR"


def etiketle(ogr_siralama, taban, alt_limit):
    try:
        ogr_siralama = int(ogr_siralama)
        taban = int(taban)
    except (TypeError, ValueError):
        return "unknown"

    if taban >= ogr_siralama:
        return "eligible"
    if taban >= alt_limit:
        return "risky"
    return "out"


def find_text_series(columns, candidates):
    normalized_columns = {str(column).strip(): column for column in columns}
    for candidate in candidates:
        for actual_name, original_name in normalized_columns.items():
            if actual_name.lower().replace(" ", "") == candidate.lower().replace(" ", ""):
                return normalized_columns[actual_name]
    return None


def prepare_dataframe(df):
    normalized = df.copy()
    normalized.columns = [str(column).strip() for column in normalized.columns]

    def text_series(column_name):
        if column_name in normalized.columns:
            return normalized[column_name].fillna("").astype(str).str.strip()
        return pd.Series([""] * len(normalized), index=normalized.index, dtype="object")

    def find_series(candidates):
        column_name = find_text_series(normalized.columns, candidates)
        if column_name is not None:
            return normalized[column_name].fillna("").astype(str).str.strip()
        return pd.Series([""] * len(normalized), index=normalized.index, dtype="object")

    program_adi = text_series("Program Ad\u0131")
    burs_orani = text_series("Burs/\u0130ndirim")
    puan_turu = text_series("Puan T\u00fcr\u00fc").str.upper()
    en_dusuk_siralama_raw = text_series("En D\u00fc\u015f\u00fck S\u0131ralama")
    dil = find_series(["Dil", "Dil Bilgisi", "Program Dili"])
    kontenjan = find_series(["Genel Kont.", "Genel Kontenjan", "Kontenjan", "Kont.", "Kota"])

    normalized["__program_adi"] = program_adi
    normalized["__program_adi_lower"] = program_adi.apply(turkish_lower)
    normalized["__burs_orani"] = [
        infer_burs_orani(burs_orani.iloc[index], program_adi.iloc[index])
        for index in range(len(normalized))
    ]
    normalized["__puan_turu"] = puan_turu
    normalized["__taban_siralama_raw"] = en_dusuk_siralama_raw
    normalized["__taban_siralama_numeric"] = en_dusuk_siralama_raw.apply(temizle_sayi_opsiyonel)
    normalized["__ucret_formatted"] = text_series("\u00dccret").apply(format_ucret)
    normalized["__dil"] = [
        detect_language(program_adi.iloc[index], dil.iloc[index])
        for index in range(len(normalized))
    ]
    normalized["__kontenjan"] = kontenjan
    return normalized


def get_dataset():
    absolute_path, display_path = resolve_active_data_file()
    cache_key = (str(absolute_path), absolute_path.stat().st_mtime_ns)

    with _dataset_lock:
        if _dataset_cache["key"] == cache_key and _dataset_cache["data"] is not None:
            return _dataset_cache["data"], display_path

    dataframe = pd.read_excel(absolute_path)
    dataframe = prepare_dataframe(dataframe)

    with _dataset_lock:
        _dataset_cache["key"] = cache_key
        _dataset_cache["data"] = dataframe
        _dataset_cache["path"] = display_path
        _dataset_cache["loaded_at"] = utcnow_iso()

    return dataframe, display_path


def sanitize_eklenenler(raw_items):
    cleaned_items = []
    if not isinstance(raw_items, list):
        return cleaned_items

    for item in raw_items[:MAX_PARAMETER_COUNT]:
        if not isinstance(item, dict):
            continue
        puan = str(item.get("puan", "")).strip()
        tur = str(item.get("tur", "")).strip().upper()
        sinir = str(item.get("sinir", "")).strip()
        if not puan or not tur or sinir == "":
            continue
        if temizle_sayi(puan) <= 0:
            continue
        cleaned_items.append({"puan": puan, "tur": tur, "sinir": sinir})

    return cleaned_items


def build_ranking_summary(items):
    rankings = []
    for item in items:
        ranking = str(item.get("puan", "")).strip()
        if ranking and ranking not in rankings:
            rankings.append(ranking)
    return ", ".join(rankings)


def build_score_types_summary(items):
    score_types = []
    for item in items:
        score_type = str(item.get("tur", "")).strip().upper()
        if score_type and score_type not in score_types:
            score_types.append(score_type)
    return ", ".join(score_types)


def build_result_row(row, parameter, ogr_siralama_int, alt_limit, lang):
    taban_siralama_numeric = row.get("__taban_siralama_numeric")
    status_key = etiketle(ogr_siralama_int, taban_siralama_numeric, alt_limit)
    return {
        "bolum_adi": row.get("__program_adi", ""),
        "puan_turu": row.get("Puan T\u00fcr\u00fc", ""),
        "burs_orani": translate_burs_orani(row.get("__burs_orani", ""), lang),
        "taban_siralama": row.get("En D\u00fc\u015f\u00fck S\u0131ralama", ""),
        "taban_puan": row.get("Taban Puan", ""),
        "tavan_puan": row.get("Tavan Puan", ""),
        "ucret": row.get("__ucret_formatted", ""),
        "dil": row.get("__dil", "TR"),
        "kontenjan": row.get("__kontenjan", ""),
        "etiket": localize_status(status_key, lang),
    }


def analiz_yap(df, eklenenler, lang):
    results = []
    seen = set()

    for parameter in eklenenler:
        ogr_siralama_int = temizle_sayi(parameter["puan"])
        sinir_int = temizle_sayi(parameter["sinir"])
        z_degeri = ogr_siralama_int - sinir_int
        alt_limit = z_degeri

        filtered = df
        if parameter["tur"]:
            filtered = filtered[filtered["__puan_turu"] == parameter["tur"]]

        siralama_numeric = filtered["__taban_siralama_numeric"]
        main_rows = filtered[siralama_numeric.notna()]
        missing_rows = filtered[siralama_numeric.isna()]

        for frame in (main_rows, missing_rows):
            for row in frame.to_dict("records"):
                unique_key = (
                    parameter["tur"],
                    parameter["puan"],
                    parameter["sinir"],
                    row.get("__program_adi", ""),
                    row.get("En D\u00fc\u015f\u00fck S\u0131ralama", ""),
                )
                if unique_key in seen:
                    continue
                seen.add(unique_key)
                results.append(build_result_row(row, parameter, ogr_siralama_int, alt_limit, lang))

    return results


def compress_results(results):
    raw_json = json.dumps(results, ensure_ascii=False).encode("utf-8")
    return sqlite3.Binary(zlib.compress(raw_json, level=6))


def decompress_results(blob):
    if not blob:
        return []
    return json.loads(zlib.decompress(blob).decode("utf-8"))


def save_analysis(
    student_email,
    student_phone,
    language,
    student_input,
    student_name,
    requested_department,
    ranking_summary,
    score_types_summary,
    params,
    results,
    source_file,
    duration_ms,
    status,
    error_message=None,
):
    analysis_id = uuid.uuid4().hex
    timestamp = utcnow_iso()
    with get_db_connection() as connection:
        connection.execute(
            """
            INSERT INTO analysis_runs(
                id, created_at, updated_at, student_email, student_phone, language,
                student_input, student_name, requested_department, ranking_summary,
                score_types_summary, params_json, result_blob, result_count, source_file,
                duration_ms, status, error_message, client_ip, user_agent
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                analysis_id,
                timestamp,
                timestamp,
                student_email,
                student_phone,
                normalize_lang(language),
                student_input,
                student_name,
                requested_department,
                ranking_summary,
                score_types_summary,
                json.dumps(params, ensure_ascii=False),
                compress_results(results) if status == "success" else None,
                len(results),
                source_file,
                duration_ms,
                status,
                error_message,
                get_client_ip(),
                get_user_agent(),
            ),
        )

    return analysis_id


def get_analysis(analysis_id):
    with get_db_connection() as connection:
        row = connection.execute("SELECT * FROM analysis_runs WHERE id = ?", (analysis_id,)).fetchone()
    return row


def record_download(analysis_id, filename, row_count):
    timestamp = utcnow_iso()
    with get_db_connection() as connection:
        connection.execute(
            "INSERT INTO download_events(analysis_id, created_at, filename, row_count, client_ip, user_agent) VALUES (?, ?, ?, ?, ?, ?)",
            (analysis_id, timestamp, filename, row_count, get_client_ip(), get_user_agent()),
        )
        connection.execute(
            "UPDATE analysis_runs SET download_count = download_count + 1, updated_at = ? WHERE id = ?",
            (timestamp, analysis_id),
        )


def build_report_context(row):
    results = decompress_results(row["result_blob"])
    params = json.loads(row["params_json"])
    lang = normalize_lang(row["language"])
    return {
        "analysis_id": row["id"],
        "adsoyad": row["student_input"],
        "student_email": row["student_email"],
        "student_phone": row["student_phone"],
        "lang": lang,
        "t": get_texts(lang),
        "eklenenler": params,
        "result": results,
        "tablo_basliklari": get_table_headers(lang),
        "veri_dosyasi_adi": row["source_file"],
        "download_url": url_for("indir_pdf", analysis_id=row["id"]),
        "ephemeral_path": None,
        "result_meta": {
            "analysis_id": row["id"],
            "created_at": row["created_at"],
            "duration_ms": row["duration_ms"],
            "row_count": row["result_count"],
            "download_count": row["download_count"],
            "source_file": row["source_file"],
        },
}


def render_analysis_template(
    student_email,
    *,
    student_phone="",
    lang="tr",
    adsoyad="",
    eklenenler=None,
    result=None,
    download_url=None,
    result_meta=None,
):
    normalized_lang = normalize_lang(lang)
    return render_template(
        "index.html",
        adsoyad=adsoyad,
        student_email=student_email,
        student_phone=student_phone,
        lang=normalized_lang,
        t=get_texts(normalized_lang),
        eklenenler=eklenenler or [],
        result=result,
        tablo_basliklari=get_table_headers(normalized_lang),
        veri_dosyasi_adi="",
        download_url=download_url,
        result_meta=result_meta,
        ephemeral_path=url_for("ephemeral_entry", lang=normalized_lang),
    )


@app.get("/indir-pdf/<analysis_id>")
def indir_pdf(analysis_id):
    row = get_analysis(analysis_id)
    if row is None or row["status"] != "success":
        abort(404)

    results = decompress_results(row["result_blob"])
    output = generate_pdf(row, results)
    if row["student_name"]:
        filename = f"{clean_filename(row['student_name'])}.pdf"
    else:
        filename = f"{clean_filename(row['student_input'] or 'rapor')}.pdf"
    record_download(analysis_id, filename, len(results))
    record_student_event(
        row["student_email"],
        "download",
        student_phone=row["student_phone"],
        language=row["language"],
        student_input=row["student_input"],
        student_name=row["student_name"],
        ranking_summary=row["ranking_summary"],
        score_types_summary=row["score_types_summary"],
        analysis_id=analysis_id,
        status="success",
        details={"filename": filename, "row_count": len(results), "format": "pdf"},
    )
    return send_file(
        output,
        mimetype="application/pdf",
        download_name=filename,
        as_attachment=True,
    )


def generate_excel(row, results):
    lang = normalize_lang(row["language"])
    texts = get_texts(lang)
    header_map = dict(get_table_headers(lang))
    header_map.pop("parametre", None)
    dataframe = pd.DataFrame(results).rename(columns=header_map)
    if "parametre" in dataframe.columns:
        dataframe = dataframe.drop(columns=["parametre"])
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sheet_name = texts["sheet_name"]
        dataframe.to_excel(writer, index=False, sheet_name=sheet_name, startrow=3)
        worksheet = writer.sheets[sheet_name]
        worksheet["A1"] = "{}: {}".format(texts["excel_student"], row["student_name"] or "")
        worksheet["A2"] = "{}: {}".format(texts["excel_department"], row["requested_department"])
        # worksheet["A3"] = "{}: {}".format(texts["excel_report"], row["id"])
        worksheet.freeze_panes = "A5"
        last_row = max(len(dataframe), 1) + 3
        last_col_index = max(len(dataframe.columns), 1)
        worksheet.auto_filter.ref = "A4:{}".format(get_column_letter(last_col_index) + str(last_row))
        for column_index, column_name in enumerate(dataframe.columns):
            width = max(len(str(column_name)), 18)
            if not dataframe.empty:
                width = min(max(width, int(dataframe[column_name].astype(str).str.len().max())), 40)
            col_letter = get_column_letter(column_index + 1)
            worksheet.column_dimensions[col_letter].width = width + 2
    output.seek(0)
    return output


def generate_pdf(row, results):
    lang = normalize_lang(row["language"])
    texts = get_texts(lang)
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
    font_name = get_pdf_font_name()
    stylesheet = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title",
        parent=stylesheet["Title"],
        fontName=font_name,
        fontSize=16,
        leading=20,
        alignment=1,
    )
    body_style = ParagraphStyle(
        "Body",
        parent=stylesheet["Normal"],
        fontName=font_name,
        fontSize=10,
        leading=14,
    )

    elements = [Paragraph(texts["page_title"], title_style), Spacer(1, 12)]
    elements.append(Paragraph(f"{texts['excel_student']}: {row['student_name'] or ''}", body_style))
    elements.append(Paragraph(f"{texts['excel_department']}: {row['requested_department']}", body_style))
    elements.append(Spacer(1, 12))

    headers = [label for _, label in get_table_headers(lang)]
    table_data = [headers]
    for item in results:
        table_data.append([str(item.get(key, "")) for key, _ in get_table_headers(lang)])

    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dbe9f4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    elements.append(table)
    doc.build(elements)
    output.seek(0)
    return output


def format_student_status(profile_row):
    if profile_row["download_count"] > 0:
        return "İndirme yaptı"
    if profile_row["analysis_count"] > 0:
        return "Sadece analiz"
    return "Sadece giriş"


def format_student_event_label(event_type):
    labels = {
        "login": "Giriş",
        "analysis": "Analiz",
        "download": "İndirme",
        "guide_open": "Kılavuz",
        "scenario_add": "Senaryo eklendi",
        "scenario_remove": "Senaryo silindi",
    }
    return labels.get(event_type, event_type)


def get_admin_metrics():
    today_prefix = datetime.now(timezone.utc).date().isoformat()
    with get_db_connection() as connection:
        recent_students = []
        for row in connection.execute(
            """
            SELECT student_email, student_phone, preferred_language, last_student_input,
                   last_student_name, last_ranking_summary, last_score_types_summary,
                   login_count, analysis_count, download_count, guide_count, event_count,
                   first_seen_at, last_seen_at, last_event_type, last_analysis_id
            FROM student_profiles
            ORDER BY last_seen_at DESC
            LIMIT 100
            """
        ).fetchall():
            item = dict(row)
            item["status_label"] = format_student_status(row)
            item["event_label"] = format_student_event_label(row["last_event_type"])
            recent_students.append(item)

        recent_student_events = []
        for row in connection.execute(
            """
            SELECT created_at, student_email, student_phone, language, event_type, student_input,
                   student_name, ranking_summary, score_types_summary, analysis_id, status
            FROM student_events
            ORDER BY created_at DESC
            LIMIT 150
            """
        ).fetchall():
            item = dict(row)
            item["event_label"] = format_student_event_label(row["event_type"])
            recent_student_events.append(item)

        metrics = {
            "total_students": connection.execute("SELECT COUNT(*) FROM student_profiles").fetchone()[0],
            "total_logins": connection.execute("SELECT COALESCE(SUM(login_count), 0) FROM student_profiles").fetchone()[0],
            "total_analyses": connection.execute(
                "SELECT COUNT(*) FROM analysis_runs WHERE status = 'success'"
            ).fetchone()[0],
            "analyses_today": connection.execute(
                "SELECT COUNT(*) FROM analysis_runs WHERE status = 'success' AND created_at LIKE ?",
                (today_prefix + "%",),
            ).fetchone()[0],
            "total_downloads": connection.execute("SELECT COUNT(*) FROM download_events").fetchone()[0],
            "total_errors": connection.execute(
                "SELECT COUNT(*) FROM analysis_runs WHERE status = 'error'"
            ).fetchone()[0],
            "avg_duration_ms": connection.execute(
                "SELECT COALESCE(AVG(duration_ms), 0) FROM analysis_runs WHERE status = 'success'"
            ).fetchone()[0],
            "recent_analyses": connection.execute(
                "SELECT id, created_at, student_email, student_phone, student_name, student_input, ranking_summary, score_types_summary, result_count, duration_ms, source_file, status, download_count FROM analysis_runs ORDER BY created_at DESC LIMIT 25"
            ).fetchall(),
            "recent_downloads": connection.execute(
                "SELECT download_events.analysis_id, download_events.created_at, download_events.filename, download_events.row_count, analysis_runs.student_email, analysis_runs.student_name, analysis_runs.ranking_summary FROM download_events LEFT JOIN analysis_runs ON analysis_runs.id = download_events.analysis_id ORDER BY download_events.created_at DESC LIMIT 25"
            ).fetchall(),
            "recent_logs": connection.execute(
                "SELECT created_at, level, event_type, message FROM app_logs ORDER BY created_at DESC LIMIT 50"
            ).fetchall(),
            "recent_students": recent_students,
            "recent_student_events": recent_student_events,
        }
    return metrics


def get_cache_status():
    with _dataset_lock:
        cached_data = _dataset_cache["data"]
        return {
            "path": _dataset_cache["path"],
            "loaded_at": _dataset_cache["loaded_at"],
            "row_count": int(len(cached_data)) if cached_data is not None else 0,
        }


@app.before_request
def before_request():
    g.request_started_at = time.perf_counter()
    maybe_cleanup()


def _handle_student_login(lang):
    texts = get_texts(lang)
    student_email = request.form.get("email", "").strip().lower()
    student_phone = normalize_phone(request.form.get("phone", ""))
    privacy_consent = request.form.get("privacy_consent") == "on"
    info_consent = request.form.get("info_consent") == "on"
    if not is_valid_email(student_email):
        flash(texts["invalid_email"], "danger")
        return render_template(
            "student_login.html",
            lang=lang,
            t=texts,
            email=student_email,
            phone=student_phone,
            privacy_consent=privacy_consent,
            info_consent=info_consent,
        )
    if not student_phone:
        flash(texts["invalid_phone"], "danger")
        return render_template(
            "student_login.html",
            lang=lang,
            t=texts,
            email=student_email,
            phone=student_phone,
            privacy_consent=privacy_consent,
            info_consent=info_consent,
        )
    if not privacy_consent:
        flash(texts["invalid_privacy_consent"], "danger")
        return render_template(
            "student_login.html",
            lang=lang,
            t=texts,
            email=student_email,
            phone=student_phone,
            privacy_consent=privacy_consent,
            info_consent=info_consent,
        )
    if not info_consent:
        flash(texts["invalid_info_consent"], "danger")
        return render_template(
            "student_login.html",
            lang=lang,
            t=texts,
            email=student_email,
            phone=student_phone,
            privacy_consent=privacy_consent,
            info_consent=info_consent,
        )

    record_student_event(
        student_email,
        "login",
        student_phone=student_phone,
        language=lang,
        details={"ip": get_client_ip(), "privacy_consent": True},
    )
    log_event(
        "INFO",
        "student_login",
        "Ogrenci girisi alindi.",
        {"student_email": student_email, "student_phone": student_phone, "language": lang},
    )
    return render_analysis_template(student_email=student_email, student_phone=student_phone, lang=lang)


@app.route("/", methods=["GET", "POST"])
def index():
    lang = normalize_lang(request.args.get("lang") or request.form.get("lang"))
    if request.method == "POST":
        return _handle_student_login(lang)
    texts = get_texts(lang)
    return render_template(
        "student_login.html",
        lang=lang,
        t=texts,
        email="",
        phone="",
        privacy_consent=False,
        info_consent=False,
    )


@app.post("/analiz")
def analyze():
    student_email = request.form.get("student_email", "").strip().lower()
    student_phone = normalize_phone(request.form.get("student_phone", ""))
    lang = normalize_lang(request.form.get("lang"))
    texts = get_texts(lang)
    adsoyad_ve_bolum = request.form.get("adsoyad", "").strip()

    if not is_valid_email(student_email) or not student_phone:
        flash(texts["invalid_session"], "warning")
        return redirect(url_for("student_login", lang=lang))

    try:
        raw_eklenenler = json.loads(request.form.get("eklenenler", "[]"))
    except json.JSONDecodeError:
        flash(texts["invalid_scenarios"], "danger")
        return render_analysis_template(
            student_email=student_email,
            student_phone=student_phone,
            lang=lang,
            adsoyad=adsoyad_ve_bolum,
        )

    eklenenler = sanitize_eklenenler(raw_eklenenler)
    if not eklenenler:
        flash(texts["no_scenario"], "warning")
        return render_analysis_template(
            student_email=student_email,
            student_phone=student_phone,
            lang=lang,
            adsoyad=adsoyad_ve_bolum,
            eklenenler=eklenenler,
        )

    if len(raw_eklenenler) > MAX_PARAMETER_COUNT:
        flash(
            texts["too_many_scenarios"].format(max_count=MAX_PARAMETER_COUNT),
            "warning",
        )

    if "," in adsoyad_ve_bolum:
        student_name, requested_department = [item.strip() for item in adsoyad_ve_bolum.split(",", 1)]
    else:
        student_name = adsoyad_ve_bolum
        requested_department = ""
    ranking_summary = build_ranking_summary(eklenenler)
    score_types_summary = build_score_types_summary(eklenenler)
    active_data_path = get_active_data_file_setting()

    started_at = time.perf_counter()
    try:
        dataframe, source_file = get_dataset()
        results = analiz_yap(dataframe, eklenenler, lang)
        duration_ms = int((time.perf_counter() - started_at) * 1000)
        analysis_id = save_analysis(
            student_email=student_email,
            student_phone=student_phone,
            language=lang,
            student_input=adsoyad_ve_bolum,
            student_name=student_name,
            requested_department=requested_department,
            ranking_summary=ranking_summary,
            score_types_summary=score_types_summary,
            params=eklenenler,
            results=results,
            source_file=source_file,
            duration_ms=duration_ms,
            status="success",
        )
        record_student_event(
            student_email,
            "analysis",
            student_phone=student_phone,
            language=lang,
            student_input=adsoyad_ve_bolum,
            student_name=student_name,
            ranking_summary=ranking_summary,
            score_types_summary=score_types_summary,
            analysis_id=analysis_id,
            status="success",
            details={"result_count": len(results), "source_file": source_file},
        )
        log_event(
            "INFO",
            "analysis_success",
            "Analiz tamamlandi.",
            {
                "analysis_id": analysis_id,
                "student_email": student_email,
                "student_name": student_name,
                "ranking_summary": ranking_summary,
                "result_count": len(results),
                "duration_ms": duration_ms,
                "source_file": source_file,
            },
        )
        if not results:
            flash(texts["no_result_flash"], "warning")
        return render_analysis_template(
            student_email=student_email,
            student_phone=student_phone,
            lang=lang,
            adsoyad=adsoyad_ve_bolum,
            eklenenler=eklenenler,
            result=results,
            download_url=url_for("indir_pdf", analysis_id=analysis_id),
            result_meta={
                "analysis_id": analysis_id,
                "created_at": utcnow_iso(),
                "duration_ms": duration_ms,
                "row_count": len(results),
                "download_count": 0,
                "source_file": source_file,
            },
        )
    except Exception as exc:
        duration_ms = int((time.perf_counter() - started_at) * 1000)
        try:
            source_file = resolve_active_data_file()[1]
        except FileNotFoundError:
            source_file = active_data_path
        analysis_id = save_analysis(
            student_email=student_email,
            student_phone=student_phone,
            language=lang,
            student_input=adsoyad_ve_bolum,
            student_name=student_name,
            requested_department=requested_department,
            ranking_summary=ranking_summary,
            score_types_summary=score_types_summary,
            params=eklenenler,
            results=[],
            source_file=source_file,
            duration_ms=duration_ms,
            status="error",
            error_message=str(exc),
        )
        record_student_event(
            student_email,
            "analysis",
            student_phone=student_phone,
            language=lang,
            student_input=adsoyad_ve_bolum,
            student_name=student_name,
            ranking_summary=ranking_summary,
            score_types_summary=score_types_summary,
            analysis_id=analysis_id,
            status="error",
            details={"error": str(exc)},
        )
        log_event(
            "ERROR",
            "analysis_error",
            "Analiz hata ile sonlandi.",
            {"analysis_id": analysis_id, "student_email": student_email, "error": str(exc)},
        )
        flash(texts["data_file_error"].format(error=exc), "danger")
        return render_analysis_template(
            student_email=student_email,
            student_phone=student_phone,
            lang=lang,
            adsoyad=adsoyad_ve_bolum,
            eklenenler=eklenenler,
        )


@app.get("/analiz")
def analyze_redirect():
    return redirect(url_for("index", lang=normalize_lang(request.args.get("lang"))))


@app.get("/oturum-gecici")
def ephemeral_entry():
    return redirect(url_for("index", lang=normalize_lang(request.args.get("lang"))))


@app.route("/giris", methods=["GET", "POST"])
def student_login():
    return redirect(url_for("index", lang=normalize_lang(request.args.get("lang") or request.form.get("lang"))))


@app.get("/cikis")
def student_logout():
    return redirect(url_for("index", lang=normalize_lang(request.args.get("lang"))))


@app.get("/rapor/<analysis_id>")
def rapor(analysis_id):
    return redirect(url_for("index", lang=normalize_lang(request.args.get("lang"))))


@app.get("/admin/rapor/<analysis_id>")
@admin_required
def admin_report(analysis_id):
    row = get_analysis(analysis_id)
    if row is None:
        abort(404)
    if row["status"] != "success":
        flash("Bu rapor olusturulamadi.", "danger")
        return redirect(url_for("admin_dashboard"))
    return render_template("index.html", **build_report_context(row))


@app.get("/indir/<analysis_id>")
def indir(analysis_id):
    row = get_analysis(analysis_id)
    if row is None or row["status"] != "success":
        abort(404)

    results = decompress_results(row["result_blob"])
    output = generate_excel(row, results)
    if row["student_name"]:
        file_base = clean_filename(row["student_name"])
        filename = f"{file_base}.xlsx"
    else:
        file_base = clean_filename(row["student_input"] or "rapor")
        filename = f"{file_base}.xlsx"
    record_download(analysis_id, filename, len(results))
    record_student_event(
        row["student_email"],
        "download",
        student_phone=row["student_phone"],
        language=row["language"],
        student_input=row["student_input"],
        student_name=row["student_name"],
        ranking_summary=row["ranking_summary"],
        score_types_summary=row["score_types_summary"],
        analysis_id=analysis_id,
        status="success",
        details={"filename": filename, "row_count": len(results)},
    )
    log_event(
        "INFO",
        "download_success",
        "Excel indirildi.",
        {
            "analysis_id": analysis_id,
            "student_email": row["student_email"],
            "filename": filename,
            "row_count": len(results),
        },
    )
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.post("/ogrenci-olay")
def ogrenci_olay():
    payload = request.get_json(silent=True) or request.form
    student_email = str(payload.get("student_email", "")).strip().lower()
    student_phone = normalize_phone(payload.get("student_phone", ""))
    lang = normalize_lang(payload.get("lang"))
    event_type = str(payload.get("event_type", "")).strip()
    if not is_valid_email(student_email) or not event_type:
        return ("", 204)

    record_student_event(
        student_email,
        event_type,
        student_phone=student_phone,
        language=lang,
        student_input=str(payload.get("student_input", "")).strip(),
        student_name=str(payload.get("student_name", "")).strip(),
        ranking_summary=str(payload.get("ranking_summary", "")).strip(),
        score_types_summary=str(payload.get("score_types_summary", "")).strip(),
        analysis_id=str(payload.get("analysis_id", "")).strip(),
        status=str(payload.get("status", "success")).strip() or "success",
        details=payload.get("details", {}),
    )
    return {"status": "ok"}


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        username = request.form.get("username", "")
        password = request.form.get("password", "")

        if not admin_credentials_configured():
            flash(
                "Admin paneli icin Render ortam degiskenlerinde ADMIN_USERNAME ve ADMIN_PASSWORD tanimlayin.",
                "warning",
            )
            return render_template("admin_login.html")

        if verify_admin_credentials(username, password):
            session["is_admin"] = True
            log_event("INFO", "admin_login_success", "Admin girisi basarili.", {"username": username})
            target = request.args.get("next") or url_for("admin_dashboard")
            return redirect(target)

        log_event(
            "WARNING",
            "admin_login_failed",
            "Admin girisi reddedildi.",
            {"username": username, "ip": get_client_ip()},
        )
        flash("Giris bilgileri hatali.", "danger")

    return render_template("admin_login.html")


@app.get("/admin/logout")
def admin_logout():
    session.pop("is_admin", None)
    flash("Admin oturumu kapatildi.", "info")
    return redirect(url_for("admin_login"))


@app.route("/admin", methods=["GET", "POST"])
@admin_required
def admin_dashboard():
    if request.method == "POST":
        action = request.form.get("action", "")
        if action == "set_data_file":
            selected_file = request.form.get("active_data_file", "")
            available_files = list_available_data_files()
            if selected_file not in available_files:
                flash("Gecersiz veri dosyasi secildi.", "danger")
            else:
                set_setting("active_data_file", selected_file)
                clear_dataset_cache()
                log_event(
                    "INFO",
                    "active_data_file_changed",
                    "Aktif veri dosyasi guncellendi.",
                    {"active_data_file": selected_file},
                )
                flash("Aktif veri dosyasi guncellendi.", "success")
        elif action == "refresh_cache":
            clear_dataset_cache()
            try:
                get_dataset()
                flash("Veri cache'i yenilendi.", "success")
            except Exception as exc:
                flash("Cache yenilenemedi: {}".format(exc), "danger")
        elif action == "cleanup":
            maybe_cleanup(force=True)
            flash("Eski log ve rapor kayitlari temizlendi.", "success")
        elif action == "purge_all_data":
            with get_db_connection() as connection:
                connection.execute("DELETE FROM download_events")
                connection.execute("DELETE FROM analysis_runs")
                connection.execute("DELETE FROM app_logs")
                connection.execute("DELETE FROM student_events")
                connection.execute("DELETE FROM student_profiles")
            flash("Tum log ve rapor verileri silindi.", "success")

    metrics = get_admin_metrics()
    cache_status = get_cache_status()
    try:
        _, active_data_file = resolve_active_data_file()
    except FileNotFoundError:
        active_data_file = get_active_data_file_setting()

    return render_template(
        "admin_dashboard.html",
        metrics=metrics,
        cache_status=cache_status,
        active_data_file=active_data_file,
        available_data_files=list_available_data_files(),
        admin_credentials_ready=admin_credentials_configured(),
        report_retention_days=REPORT_RETENTION_DAYS,
        log_retention_days=LOG_RETENTION_DAYS,
        database_path=str(DATABASE_PATH),
    )


@app.get("/kullanimklavuzu")
def kullanim_klavuzu():
    lang = normalize_lang(request.args.get("lang"))
    student_email = request.args.get("student_email", "").strip().lower()
    student_phone = normalize_phone(request.args.get("student_phone", ""))
    if is_valid_email(student_email):
        record_student_event(
            student_email,
            "guide_open",
            student_phone=student_phone,
            language=lang,
            status="success",
        )
    kilavuz_yolu = APP_ROOT / ("kullanimklavuzu_en.txt" if lang == "en" else "kullanimklavuzu.txt")
    return send_file(str(kilavuz_yolu), mimetype="text/plain; charset=utf-8")


@app.get("/health")
def health():
    try:
        dataframe, source_file = get_dataset()
        return {
            "status": "ok",
            "active_data_file": source_file,
            "rows": int(len(dataframe)),
            "database": str(DATABASE_PATH),
        }
    except Exception as exc:
        return {"status": "error", "message": str(exc)}, 500


init_db()
maybe_cleanup(force=True)

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
